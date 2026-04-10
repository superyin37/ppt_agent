from __future__ import annotations

import hashlib
import json
import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile

from sqlalchemy.orm import Session

from db.models.asset import Asset
from db.models.material_item import MaterialItem
from db.models.material_package import MaterialPackage
from db.models.project import Project, ProjectBrief
from schema.common import AssetType, ProjectStatus

logger = logging.getLogger(__name__)

TEXT_EXTENSIONS = {".md", ".txt", ".json", ".html", ".htm"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".svg"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".csv"}
CHART_VARIANT_EXTENSIONS = {".json", ".svg", ".html"}


def infer_logical_key(path: Path) -> str:
    name = path.name

    match = re.search(r"参考案例(\d+)_图片_\d+_", name)
    if match:
        return f"reference.case.{match.group(1)}.images"

    match = re.search(r"参考案例(\d+)_缩略图", name)
    if match:
        return f"reference.case.{match.group(1)}.thumbnail"

    match = re.search(r"参考案例(\d+)_archdaily", name)
    if match:
        return f"reference.case.{match.group(1)}.source"

    match = re.search(r"案例(\d+)_评价和分析", name)
    if match:
        return f"reference.case.{match.group(1)}.analysis"

    match = re.search(r"经济背景 - ([^_]+)_chart_(\d+)_", name)
    if match:
        topic_map = {
            "城市经济": "city",
            "产业发展": "industry",
            "消费水平": "consumption",
        }
        return f"economy.{topic_map.get(match.group(1), 'general')}.chart.{match.group(2)}"

    static_mapping = {
        "场地四至分析": "site.boundary.image",
        "场地poi": "site.poi.table",
        "场地坐标": "site.coordinate.text",
        "外部交通站点_POI": "site.transport.station.table",
        "外部交通站点": "site.transport.station.image",
        "外部交通_POI": "site.transport.external.table",
        "外部交通": "site.transport.external.image",
        "枢纽站点_POI": "site.transport.hub.table",
        "枢纽站点": "site.transport.hub.image",
        "周边基础设施建设规划_POI": "site.infrastructure.plan.table",
        "周边基础设施建设规划": "site.infrastructure.plan.image",
        "区域开发情况_POI": "site.development.table",
        "区域开发情况": "site.development.image",
        "附近同类型产品分析_POI": "site.competitor.table",
        "附近同类型产品分析": "site.competitor.image",
        "设计建议书大纲": "brief.design_outline",
        "manus提示词": "brief.manus_prompt",
    }
    for prefix, logical_key in static_mapping.items():
        if name.startswith(prefix):
            return logical_key

    return f"misc.{path.suffix.lstrip('.').lower()}.{path.stem}"


def _guess_kind(path: Path, logical_key: str) -> str:
    suffix = path.suffix.lower()
    if logical_key.startswith("economy.") and "_chart_" in path.stem:
        return "chart_bundle" if suffix in CHART_VARIANT_EXTENSIONS else "chart"
    if suffix in IMAGE_EXTENSIONS:
        return "image"
    if suffix in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"
    if suffix in TEXT_EXTENSIONS:
        return "document"
    return "binary"


def _read_text_excerpt(path: Path, max_chars: int = 2000) -> str | None:
    if path.suffix.lower() not in TEXT_EXTENSIONS:
        return None
    try:
        return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]
    except Exception:
        return None


def _read_json_payload(path: Path) -> dict | None:
    if path.suffix.lower() != ".json":
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    except Exception:
        return None


def _extract_xlsx_metadata(path: Path) -> dict:
    metadata: dict = {"file_size": path.stat().st_size}
    try:
        with ZipFile(path) as zf:
            metadata["zip_entries"] = len(zf.namelist())
    except Exception:
        pass
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        metadata["sheet_names"] = wb.sheetnames
        preview_rows = []
        for ws in wb.worksheets[:2]:
            rows = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append([cell for cell in row[:8]])
                if idx >= 4:
                    break
            preview_rows.append({"sheet": ws.title, "rows": rows})
        metadata["preview_rows"] = preview_rows
        wb.close()
    except Exception:
        metadata.setdefault("sheet_names", [])
    return metadata


def _hash_path(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.name.encode("utf-8", errors="ignore"))
    digest.update(str(path.stat().st_size).encode())
    return digest.hexdigest()


def _group_chart_variants(paths: Iterable[Path]) -> tuple[list[Path], list[tuple[Path, dict[str, Path]]]]:
    grouped: dict[str, dict[str, Path]] = defaultdict(dict)
    regular: list[Path] = []
    for path in paths:
        if "_chart_" in path.stem and path.suffix.lower() in CHART_VARIANT_EXTENSIONS:
            grouped[path.stem][path.suffix.lstrip(".").lower()] = path
        else:
            regular.append(path)

    chart_groups: list[tuple[Path, dict[str, Path]]] = []
    for variants in grouped.values():
        base = variants.get("svg") or variants.get("json") or variants.get("html")
        if base:
            chart_groups.append((base, variants))
    return regular, chart_groups


def _build_item_payload(path: Path, logical_key: str, variants: dict[str, Path] | None = None) -> dict:
    kind = _guess_kind(path, logical_key)
    metadata_json = {"basename": path.stem, "file_name": path.name}
    preview_url = None
    content_url = None
    text_content = None
    structured_data = None

    if variants:
        metadata_json["variants"] = {k: str(v) for k, v in variants.items()}
        preview_candidate = variants.get("svg")
        content_candidate = variants.get("html") or preview_candidate or variants.get("json")
        preview_url = preview_candidate.resolve().as_uri() if preview_candidate else None
        content_url = content_candidate.resolve().as_uri() if content_candidate else None
        if variants.get("json"):
            structured_data = _read_json_payload(variants["json"])
        kind = "chart_bundle"
        file_format = "bundle"
    else:
        suffix = path.suffix.lower()
        file_format = suffix.lstrip(".")
        if suffix in IMAGE_EXTENSIONS:
            preview_url = path.resolve().as_uri()
            content_url = preview_url
        elif suffix in TEXT_EXTENSIONS:
            text_content = _read_text_excerpt(path)
            content_url = path.resolve().as_uri()
            if suffix == ".json":
                structured_data = _read_json_payload(path)
        elif suffix in SPREADSHEET_EXTENSIONS:
            structured_data = _extract_xlsx_metadata(path)
            content_url = path.resolve().as_uri()
        else:
            content_url = path.resolve().as_uri()

    return {
        "logical_key": logical_key,
        "kind": kind,
        "format": file_format,
        "title": path.stem,
        "source_path": str(path.resolve()),
        "preview_url": preview_url,
        "content_url": content_url,
        "text_content": text_content,
        "structured_data": structured_data,
        "tags": [logical_key.split(".")[0]],
        "source_hash": _hash_path(path),
        "metadata_json": metadata_json,
    }


def build_manifest(items: list[MaterialItem]) -> dict:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        grouped[item.logical_key].append({
            "id": str(item.id),
            "kind": item.kind,
            "format": item.format,
            "title": item.title,
            "source_path": item.source_path,
            "preview_url": item.preview_url,
            "content_url": item.content_url,
        })
    return {
        "logical_keys": sorted(grouped.keys()),
        "items": [{"logical_key": key, "entries": entries} for key, entries in sorted(grouped.items())],
    }


def build_summary(items: list[MaterialItem]) -> dict:
    logical_key_counts: dict[str, int] = defaultdict(int)
    evidence_snippets: list[dict] = []
    case_numbers = set()
    for item in items:
        logical_key_counts[item.logical_key] += 1
        if item.logical_key.startswith("reference.case."):
            case_numbers.add(item.logical_key.split(".")[2])
        if item.text_content:
            evidence_snippets.append({
                "logical_key": item.logical_key,
                "title": item.title,
                "snippet": item.text_content[:240],
            })

    return {
        "item_count": len(items),
        "logical_key_counts": dict(sorted(logical_key_counts.items())),
        "evidence_snippets": evidence_snippets[:20],
        "case_count": len(case_numbers),
        "chart_count": sum(1 for item in items if item.kind == "chart_bundle"),
    }


def ensure_project_brief_from_package(project_id, project_name: str, summary_json: dict, db: Session, items: list[MaterialItem] | None = None) -> ProjectBrief:
    latest = (
        db.query(ProjectBrief)
        .filter(ProjectBrief.project_id == project_id)
        .order_by(ProjectBrief.version.desc())
        .first()
    )
    if latest:
        return latest

    evidence_text = " ".join(snippet["snippet"] for snippet in summary_json.get("evidence_snippets", []))

    # --- 从设计建议书大纲中提取结构化信息 ---
    design_outline_text = ""
    if items:
        for item in items:
            if item.logical_key == "brief.design_outline" and item.text_content:
                design_outline_text = item.text_content
                break

    city, province, district, site_address = _extract_location(design_outline_text)
    building_type = _detect_building_type(evidence_text, design_outline_text)
    style_prefs = _extract_style_preferences(design_outline_text, evidence_text)
    far = _extract_far(design_outline_text)

    brief = ProjectBrief(
        project_id=project_id,
        version=1,
        status="confirmed",
        building_type=building_type,
        client_name=project_name,
        style_preferences=style_prefs,
        city=city,
        province=province,
        district=district,
        site_address=site_address,
        far=far,
        missing_fields=[],
    )
    db.add(brief)
    db.flush()
    return brief


def _extract_location(text: str) -> tuple[str | None, str | None, str | None, str | None]:
    """从设计建议书文本中提取城市、省份、区县、地址。"""
    if not text:
        return None, None, None, None

    city = province = district = site_address = None

    # 匹配 "XX市XX区XX" 格式的项目标题
    m = re.search(r"([\u4e00-\u9fff]{2,6}[市州])([\u4e00-\u9fff]{1,6}[区县市旗])", text[:500])
    if m:
        city = m.group(1)
        district = m.group(2)

    # 省份推断
    province_map = {
        "武汉": "湖北省", "十堰": "湖北省", "宜昌": "湖北省", "襄阳": "湖北省", "荆州": "湖北省",
        "北京": "北京市", "上海": "上海市", "天津": "天津市", "重庆": "重庆市",
        "广州": "广东省", "深圳": "广东省", "东莞": "广东省", "佛山": "广东省",
        "杭州": "浙江省", "宁波": "浙江省", "温州": "浙江省",
        "南京": "江苏省", "苏州": "江苏省", "无锡": "江苏省",
        "成都": "四川省", "长沙": "湖南省", "郑州": "河南省",
        "西安": "陕西省", "济南": "山东省", "青岛": "山东省",
        "福州": "福建省", "厦门": "福建省", "合肥": "安徽省",
        "南昌": "江西省", "昆明": "云南省", "贵阳": "贵州省",
        "南宁": "广西", "海口": "海南省", "太原": "山西省",
        "石家庄": "河北省", "哈尔滨": "黑龙江省", "长春": "吉林省", "沈阳": "辽宁省",
        "呼和浩特": "内蒙古", "兰州": "甘肃省", "银川": "宁夏", "西宁": "青海省",
        "乌鲁木齐": "新疆", "拉萨": "西藏",
    }
    if city:
        city_short = city.rstrip("市州")
        province = province_map.get(city_short)

    # 尝试提取更具体的地址
    addr_m = re.search(r"([\u4e00-\u9fff]{2,6}[市州][\u4e00-\u9fff]{1,6}[区县市旗][\u4e00-\u9fff]{2,20})", text[:500])
    if addr_m:
        site_address = addr_m.group(1)

    return city, province, district, site_address


_BUILDING_TYPE_KEYWORDS: dict[str, list[str]] = {
    "public": ["公厕", "公共厕所", "公共卫生间"],
    "cultural": ["文化", "博物", "展览", "美术馆", "图书馆", "文化中心"],
    "education": ["学校", "教育", "大学", "中学", "小学", "幼儿园", "校园"],
    "office": ["办公", "写字楼", "总部", "企业"],
    "residential": ["住宅", "居住", "公寓", "小区", "楼盘"],
    "commercial": ["商业", "购物", "商场", "零售", "商圈"],
    "hotel": ["酒店", "宾馆", "度假村", "民宿"],
    "healthcare": ["医院", "医疗", "诊所", "康养"],
    "sports": ["体育", "运动", "体育馆", "球场"],
    "mixed": ["综合体", "综合", "混合"],
}


def _detect_building_type(evidence_text: str, design_outline_text: str) -> str:
    """从素材文本中检测建筑类型。"""
    combined = (design_outline_text[:1000] + " " + evidence_text)[:3000]
    best_type = "mixed"
    best_score = 0
    for btype, keywords in _BUILDING_TYPE_KEYWORDS.items():
        score = sum(combined.count(kw) for kw in keywords)
        if score > best_score:
            best_score = score
            best_type = btype
    return best_type


def _extract_style_preferences(design_outline_text: str, evidence_text: str) -> list[str]:
    """从文本中提取风格关键词。"""
    combined = (design_outline_text + " " + evidence_text)[:5000]
    style_keywords = [
        "现代简约", "现代", "简约", "中式", "新中式", "禅意", "工业风", "工业",
        "极简", "参数化", "自然", "生态", "科技", "智慧", "人文",
        "传统", "古典", "地域", "文化", "绿色", "可持续",
    ]
    found = []
    for kw in style_keywords:
        if kw in combined:
            # 避免子串重复（如 "现代简约" 已包含 "现代" 和 "简约"）
            if not any(kw != f and kw in f for f in found):
                found.append(kw)
    return found[:6] if found else ["modern", "minimal"]


def _extract_far(text: str) -> float | None:
    """从文本中提取容积率。"""
    if not text:
        return None
    m = re.search(r"容积率[：:≤≥\s]*([0-9]+\.?[0-9]*)", text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


def derive_assets_from_items(project_id, package_id, items: list[MaterialItem], db: Session) -> list[Asset]:
    assets: list[Asset] = []
    for item in items:
        asset_type = None
        render_role = None
        image_url = item.preview_url
        data_json = item.structured_data
        summary = item.text_content[:500] if item.text_content else None

        if item.kind == "chart_bundle":
            asset_type = AssetType.CHART.value
            render_role = "chart"
            image_url = item.preview_url or item.content_url
        elif item.kind == "image":
            asset_type = AssetType.MAP.value if item.logical_key.startswith("site.") else AssetType.IMAGE.value
            render_role = "image"
        elif item.kind == "spreadsheet":
            asset_type = AssetType.KPI_TABLE.value
            render_role = "table"
        elif item.kind == "document":
            asset_type = AssetType.TEXT_SUMMARY.value
            render_role = "summary"

        if not asset_type:
            continue

        asset = Asset(
            project_id=project_id,
            package_id=package_id,
            source_item_id=item.id,
            asset_type=asset_type,
            subtype=item.kind,
            title=item.title,
            image_url=image_url,
            data_json=data_json,
            config_json={
                "preview_url": item.preview_url,
                "content_url": item.content_url,
                "source_path": item.source_path,
                "metadata_json": item.metadata_json,
            },
            summary=summary,
            logical_key=item.logical_key,
            variant=item.format,
            render_role=render_role,
            is_primary=True,
            status="ready",
            source_info={"material_item_id": str(item.id), "logical_key": item.logical_key},
        )
        db.add(asset)
        db.flush()
        assets.append(asset)

    case_groups: dict[str, dict[str, list[MaterialItem] | MaterialItem]] = defaultdict(dict)
    for item in items:
        if not item.logical_key.startswith("reference.case."):
            continue
        case_no = item.logical_key.split(".")[2]
        suffix = item.logical_key.split(".", 3)[3]
        if suffix == "images":
            case_groups[case_no].setdefault("images", [])
            assert isinstance(case_groups[case_no]["images"], list)
            case_groups[case_no]["images"].append(item)
        else:
            case_groups[case_no][suffix] = item

    for case_no, group in case_groups.items():
        thumbnail = group.get("thumbnail")
        source = group.get("source")
        analysis = group.get("analysis")
        summary_parts = []
        for item in (analysis, source):
            if isinstance(item, MaterialItem) and item.text_content:
                summary_parts.append(item.text_content[:280])
        case_card = Asset(
            project_id=project_id,
            package_id=package_id,
            source_item_id=thumbnail.id if isinstance(thumbnail, MaterialItem) else None,
            asset_type=AssetType.CASE_CARD.value,
            subtype="case_card",
            title=f"参考案例 {case_no}",
            image_url=thumbnail.preview_url if isinstance(thumbnail, MaterialItem) else None,
            data_json={
                "case_no": case_no,
                "thumbnail_item_id": str(thumbnail.id) if isinstance(thumbnail, MaterialItem) else None,
                "source_item_id": str(source.id) if isinstance(source, MaterialItem) else None,
                "analysis_item_id": str(analysis.id) if isinstance(analysis, MaterialItem) else None,
                "image_item_ids": [str(i.id) for i in group.get("images", [])] if isinstance(group.get("images"), list) else [],
            },
            config_json={
                "thumbnail_url": thumbnail.preview_url if isinstance(thumbnail, MaterialItem) else None,
                "source_url": source.content_url if isinstance(source, MaterialItem) else None,
                "analysis_url": analysis.content_url if isinstance(analysis, MaterialItem) else None,
            },
            summary="\n".join(summary_parts)[:1000] if summary_parts else None,
            logical_key=f"reference.case.{case_no}.card",
            variant="card",
            render_role="case_card",
            is_primary=True,
            status="ready",
        )
        db.add(case_card)
        db.flush()
        assets.append(case_card)

    return assets


def ingest_local_material_package(project_id, local_path: str, db: Session) -> MaterialPackage:
    path = Path(local_path)
    if not path.exists() or not path.is_dir():
        raise ValueError(f"Material package path does not exist or is not a directory: {local_path}")

    latest = (
        db.query(MaterialPackage)
        .filter(MaterialPackage.project_id == project_id)
        .order_by(MaterialPackage.version.desc())
        .first()
    )
    new_version = (latest.version + 1) if latest else 1

    package = MaterialPackage(
        project_id=project_id,
        version=new_version,
        status="ingesting",
        created_from={"type": "local_directory", "local_path": str(path.resolve())},
    )
    db.add(package)
    db.flush()

    files = sorted(p for p in path.iterdir() if p.is_file())
    regular_files, chart_groups = _group_chart_variants(files)

    items: list[MaterialItem] = []
    for file_path in regular_files:
        payload = _build_item_payload(file_path, infer_logical_key(file_path))
        item = MaterialItem(package_id=package.id, **payload)
        db.add(item)
        db.flush()
        items.append(item)

    for base_path, variants in chart_groups:
        payload = _build_item_payload(base_path, infer_logical_key(base_path), variants=variants)
        item = MaterialItem(package_id=package.id, **payload)
        db.add(item)
        db.flush()
        items.append(item)

    package.manifest_json = build_manifest(items)
    package.summary_json = build_summary(items)
    package.source_hash = hashlib.sha256(
        "".join(sorted(item.source_hash or "" for item in items)).encode("utf-8", errors="ignore")
    ).hexdigest()
    package.status = "ready"

    derive_assets_from_items(project_id, package.id, items, db)

    project = db.get(Project, project_id)
    ensure_project_brief_from_package(
        project_id=project_id,
        project_name=project.name if project else path.name,
        summary_json=package.summary_json or {},
        db=db,
        items=items,
    )
    if project:
        project.status = ProjectStatus.MATERIAL_READY.value
        project.current_phase = "material_ready"

    return package
